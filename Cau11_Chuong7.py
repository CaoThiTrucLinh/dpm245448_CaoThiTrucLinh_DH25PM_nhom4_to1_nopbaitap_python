#Cau11
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import os

# Tên file Excel
FILE_NAME = "QLNhanVien.xlsx"

# --- I. CÁC HÀM XỬ LÝ FILE (Đọc/Ghi) ---

def khoi_tao_file_excel(file_path):
    """Khởi tạo file Excel nếu chưa tồn tại và tạo tiêu đề."""
    if not os.path.exists(file_path):
        print(f"File '{file_path}' chưa tồn tại. Đang khởi tạo...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Thiết lập tiêu đề cột
        sheet.append(['STT', 'Mã', 'Tên', 'Tuổi'])
        
        # Dữ liệu mẫu ban đầu (theo hình minh họa)
        data_mau = [
            ['NV1', 'An', 18],
            ['NV2', 'Lành', 22],
            ['NV3', 'Giải', 20],
            ['NV4', 'Thoát', 19],
            ['NV5', 'Hạnh', 25],
            ['NV6', 'Phúc', 24]
        ]
        
        for i, row_data in enumerate(data_mau):
            # STT bắt đầu từ 1
            sheet.append([i + 1] + row_data) 
            
        workbook.save(file_path)
        print("Đã tạo file và dữ liệu mẫu thành công.")
        return True
    return False

def doc_danh_sach_nhan_vien(file_path):
    """Đọc toàn bộ danh sách Nhân viên (bỏ qua hàng tiêu đề) từ file Excel."""
    nhan_vien = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # Bắt đầu đọc từ hàng thứ 2 (chỉ mục 1) để bỏ qua tiêu đề
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            # Đảm bảo hàng có đủ dữ liệu (Mã, Tên, Tuổi)
            if len(row) >= 4 and row[1].value is not None: 
                nv = {
                    'STT': row[0].value,
                    'Ma': str(row[1].value),
                    'Ten': str(row[2].value),
                    'Tuoi': int(row[3].value) if row[3].value else 0
                }
                nhan_vien.append(nv)
        
        return nhan_vien
        
    except FileNotFoundError:
        print(f"Lỗi: File '{file_path}' không tồn tại.")
        return []
    except InvalidFileException:
        print(f"Lỗi: File '{file_path}' không phải là file Excel hợp lệ.")
        return []
    except Exception as e:
        print(f"Lỗi khi đọc file: {e}")
        return []

# --- II. CÁC HÀM CHỨC NĂNG (Thêm, Sắp xếp, Hiển thị) ---

def luu_danh_sach_vao_excel(file_path, danh_sach_moi):
    """Ghi lại toàn bộ danh sách Nhân viên vào file Excel."""
    try:
        workbook = openpyxl.Workbook() # Tạo Workbook mới để ghi đè toàn bộ
        sheet = workbook.active
        
        # Ghi tiêu đề
        sheet.append(['STT', 'Mã', 'Tên', 'Tuổi'])
        
        # Ghi dữ liệu mới
        for i, nv in enumerate(danh_sach_moi):
            # Tái tạo STT để đảm bảo luôn đúng thứ tự
            sheet.append([i + 1, nv['Ma'], nv['Ten'], nv['Tuoi']])
            
        workbook.save(file_path)
        return True
        
    except Exception as e:
        print(f"Lỗi khi lưu file: {e}")
        return False

def them_nhan_vien():
    """Thêm một nhân viên mới vào danh sách và lưu lại file."""
    danh_sach = doc_danh_sach_nhan_vien(FILE_NAME)
    
    ma_nv = input("Nhập Mã Nhân viên mới (ví dụ: NV7): ").upper()
    if any(nv['Ma'] == ma_nv for nv in danh_sach):
        print("Lỗi: Mã Nhân viên đã tồn tại.")
        return

    ten_nv = input("Nhập Tên Nhân viên: ")
    try:
        tuoi_nv = int(input("Nhập Tuổi: "))
    except ValueError:
        print("Lỗi: Tuổi phải là số nguyên.")
        return

    new_nv = {'STT': len(danh_sach) + 1, 'Ma': ma_nv, 'Ten': ten_nv, 'Tuoi': tuoi_nv}
    danh_sach.append(new_nv)
    
    if luu_danh_sach_vao_excel(FILE_NAME, danh_sach):
        print(f"Đã thêm Nhân viên {ma_nv} thành công.")
    else:
        print("Thêm nhân viên thất bại do lỗi lưu file.")

def hien_thi_danh_sach(danh_sach):
    """Hiển thị danh sách Nhân viên ra màn hình."""
    if not danh_sach:
        print("Danh sách Nhân viên trống.")
        return
        
    print("\n--- DANH SÁCH NHÂN VIÊN ---")
    print(f"{'STT':<4} | {'Mã NV':<6} | {'Tên NV':<15} | {'Tuổi':<5}")
    print("-" * 35)
    for nv in danh_sach:
        print(f"{nv['STT']:<4} | {nv['Ma']:<6} | {nv['Ten']:<15} | {nv['Tuoi']:<5}")
    print("-" * 35)

def sap_xep_nhan_vien_theo_tuoi():
    """Sắp xếp Nhân viên theo Tuổi tăng dần và lưu lại file."""
    danh_sach = doc_danh_sach_nhan_vien(FILE_NAME)
    
    # Sắp xếp theo 'Tuoi' tăng dần
    danh_sach.sort(key=lambda nv: nv['Tuoi'], reverse=False)
    
    if luu_danh_sach_vao_excel(FILE_NAME, danh_sach):
        print("Đã sắp xếp Nhân viên theo Tuổi tăng dần và lưu lại file thành công.")
        hien_thi_danh_sach(danh_sach)
    else:
        print("Sắp xếp thất bại do lỗi lưu file.")

# --- III. HÀM MAIN VÀ MENU ---

def menu_nhan_vien():
    """Menu chính của phần mềm Quản Lý Nhân Viên."""
    # Đảm bảo file tồn tại trước khi chạy các chức năng
    khoi_tao_file_excel(FILE_NAME)
    
    while True:
        print("\n===== QUẢN LÝ NHÂN VIÊN (EXCEL FILE) =====")
        print("1. Thêm mới Nhân viên")
        print("2. Đọc và Hiển thị danh sách")
        print("3. Sắp xếp Nhân viên theo Tuổi (tăng dần)")
        print("0. Thoát")
        
        choice = input("Chọn chức năng (0-3): ")
        
        if choice == '1':
            them_nhan_vien()
        elif choice == '2':
            danh_sach = doc_danh_sach_nhan_vien(FILE_NAME)
            hien_thi_danh_sach(danh_sach)
        elif choice == '3':
            sap_xep_nhan_vien_theo_tuoi()
        elif choice == '0':
            print("Đã thoát chương trình Quản Lý Nhân Viên.")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng chọn lại.")

# Chạy chương trình
if __name__ == "__main__":
    menu_nhan_vien()